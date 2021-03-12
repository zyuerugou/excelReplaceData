# -*- coding: utf-8 -*-
"""
从xlsx文件中替换关键词

@author: 193334
"""

import openpyxl


def read_excel(input_file_name):
    """
    从xlsx文件中读取数据
    @parameters           input_file_name       读取的文件名
    @return               list(sheet.values)    excel文件中的数据，将turple转化成list
    """
    workbook = openpyxl.load_workbook(input_file_name)
    print(workbook)
    # 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
    print(workbook.sheetnames)
    sheet = workbook.active
    
    return list(sheet.values)


def save_excel(output_file_name, data):
    '''
    保存到xlsx
    @parameters        output_file_name      保存的文件名
                       data                  待保存的excel表数据
    '''
    workbook = openpyxl.Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    sheet = workbook.active
    
    row = len(data)
    col = len(data[0])
    for i in range(row):
        for j in range(col):
            sheet.cell(i + 1, j + 1).value = data[i][j]
    
    workbook.save(output_file_name)


def data_replace(data, index, mapping):
    '''
    替换对应列中的关键词
    @parameters        data       excel表中的数据
                       index      需要替换的列
                       mapping    替换的字符串map
    @return            result     替换完成的excel表中数据
    '''
    result = []
    for data_tmp in data:
        tmp = list(data_tmp)
        for amapping in mapping:
            for i in index:
                # 空格跳过
                if tmp[i] is not None:
                    tmp[i] = str(tmp[i]).replace(str(amapping[0]), str(amapping[1]))
        
        result.append(tuple(tmp))
    
    return result

def excel_operate(file_name, file_format, index, mapping):
    '''
    处理单个excel文件：读取内容，替换，保存
    @parameters     file_name   文件名
                    index       需要替换的列
                    mapping     替换的字符串map
    @return         True        替换成功
                    False       替换失败
    '''
    try:
        data = read_excel(file_name + file_format)
        data_result = data_replace(data, index, mapping)
        save_excel(file_name + '_1' + file_format, data_result)
        return True
    except BaseException as e:
        print(e.__str__())
        return False
    


if __name__ == '__main__':
    
    # 获取映射表
    mapping = read_excel(r'对应表.xlsx')
    
    # 获取1.xlsx，并替换
    file_name = r'1'
    status = excel_operate(file_name, '.xlsx', [2], mapping)
    if status:
        print('替换 {0} 成功'.format(file_name))
    else:
        print('替换 {0} 失败'.format(file_name))
    
    # 获取2.xlsx，并替换
    file_name = r'2'
    status = excel_operate(file_name, '.xlsx', [1, 2], mapping)
    if status:
        print('替换 {0} 成功'.format(file_name))
    else:
        print('替换 {0} 失败'.format(file_name))
    
    # 3.xlsx，并替换
    file_name = r'3'
    status = excel_operate(file_name, '.xlsx', [0, 1, 2], mapping)
    if status:
        print('替换 {0} 成功'.format(file_name))
    else:
        print('替换 {0} 失败'.format(file_name))
    